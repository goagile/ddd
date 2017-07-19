from openpyxl import load_workbook

from start_examples.po.msg_model.msg_collection import MsgCollection


class ParseResult:

    def __init__(self):
        self.ru_msg_collection = MsgCollection()
        self.en_msg_collection = MsgCollection()
        self.new_ru_msg_collection = MsgCollection()
        self.new_en_msg_collection = MsgCollection()


class RuEnExcelParser:

    @classmethod
    def parse(cls, excel_path) -> ParseResult:
        result = ParseResult()

        ws = load_workbook(excel_path).get_active_sheet()

        for row in ws.iter_rows(min_row=2):
            cls.__parse_to(result.ru_msg_collection, row, str_col=1)
            cls.__parse_to(result.en_msg_collection, row, str_col=2)
            cls.__parse_to(result.new_ru_msg_collection, row, id_col=3, str_col=4)
            cls.__parse_to(result.new_en_msg_collection, row, id_col=3, str_col=5)

        return result

    @classmethod
    def __parse_to(cls, en_msg_collection, row, id_col=0, str_col=1, path_col=6):
        id_, en = row[id_col].value, row[str_col].value
        paths = cls._split_by_new_line(row[path_col].value)
        if cls.__is_plural(id_, en):
            ids = cls._split_plural_ids_by_new_line(id_)
            en_msg_collection.add_msg_plural(*ids, strs=cls._split_by_new_line(en), paths=paths)
        else:
            en_msg_collection.add_msg(id=id_, str=en, paths=paths)

    @classmethod
    def _split_plural_ids_by_new_line(cls, ids):
        splitted = cls._split_by_new_line(ids)
        if len(splitted) < 2:
            splitted.append('')
        return splitted

    @classmethod
    def _split_by_new_line(cls, str):
        if str is None:
            return ''
        splitted = str.split('\n')
        result = cls.__ignore_empty(splitted)
        return result

    @classmethod
    def __ignore_empty(cls, splitted):
        return [s.strip() for s in splitted if s]

    @classmethod
    def __is_plural(cls, id_, str):
        return bool(cls.__has_plural_values(id_) or cls.__has_plural_values(str))

    @classmethod
    def __has_plural_values(cls, value):
        if value is None:
            return False
        return bool(len(value.split('\n')) > 1)
