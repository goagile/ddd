from openpyxl import load_workbook

from start_examples.po.msg_model.msg_collection import MsgCollection


class RuEnExcelParser:

    @classmethod
    def parse(cls, excel_path):
        wb = load_workbook(excel_path)
        ws = wb.get_active_sheet()

        ru_msg_collection = MsgCollection()
        en_msg_collection = MsgCollection()

        for row in ws.iter_rows(min_row=2):
            cls.__parse_ru_to(ru_msg_collection, row)
            cls.__parse_en_to(en_msg_collection, row)

        return ru_msg_collection, en_msg_collection

    @classmethod
    def __parse_ru_to(cls, ru_msg_collection, row):
        id_, ru = row[0].value, row[1].value
        paths = cls.split_by_new_line(row[6].value)
        if cls.is_plural(id_) or cls.is_plural(ru):
            ids = cls.split_by_new_line(id_)
            ru_msg_collection.add_msg_plural(*ids, strs=cls.split_by_new_line(ru), paths=paths)
        else:
            ru_msg_collection.add_msg(id=id_, str=ru, paths=paths)

    @classmethod
    def __parse_en_to(cls, en_msg_collection, row):
        id_, en = row[0].value, row[2].value
        paths = cls.split_by_new_line(row[6].value)
        if cls.is_plural(id_) or cls.is_plural(en):
            ids = cls.split_by_new_line(id_)
            en_msg_collection.add_msg_plural(*ids, strs=cls.split_by_new_line(en), paths=paths)
        else:
            en_msg_collection.add_msg(id=id_, str=en, paths=paths)

    @classmethod
    def split_by_new_line(cls, str):
        if str is None:
            return ''
        return str.split('\n')

    @classmethod
    def is_plural(cls, value):
        if value is None:
            return False
        return bool(len(value.split('\n')) > 1)
