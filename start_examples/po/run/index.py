import codecs

from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Alignment
from openpyxl.utils import get_column_letter

from start_examples.po.parser.po_parser import PoParser


def fill_excel(excel_path, msg_collection):
    wb = Workbook()
    ws = wb.get_active_sheet()
    alignment = Alignment(wrapText=True, vertical='top')

    for row_index, msg in enumerate(msg_collection, 2):
        if msg.is_plural:
            cell_id = ws.cell(row=row_index, column=1)
            cell_id.alignment = alignment
            cell_id.value = msg.id_plural

            cell_en = ws.cell(row=row_index, column=2)
            cell_en.alignment = alignment
            cell_en.value = join_by_new_line(msg.strs)

            cell_paths = ws.cell(row=row_index, column=3)
            cell_paths.alignment = alignment
            cell_paths.value = join_by_new_line(msg.paths)
        else:
            cell_id = ws.cell(row=row_index, column=1)
            cell_id.alignment = alignment
            cell_id.value = msg.id

            cell_en = ws.cell(row=row_index, column=2)
            cell_en.alignment = alignment
            cell_en.value = msg.str

            cell_paths = ws.cell(row=row_index, column=3)
            cell_paths.alignment = alignment
            cell_paths.value = join_by_new_line(msg.paths)

        # ws.row_dimensions[row_index].height = 15

    widths = (25, 30, 40, 20, 40)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(excel_path)


def join_by_new_line(collection):
    return '\n'.join(p for p in collection)


if __name__ == '__main__':
    excel_path = 'text.xlsx'
    path = 'test.po'
    with codecs.open(path, "r", "utf-8") as file:
        lines = file.readlines()

    po_parser = PoParser.new()

    po_parser.parse_lines(lines)

    msg_collection = po_parser.msg_collection

    # print(msg_collection)

    for m in msg_collection:
        print(m)

    fill_excel(excel_path, msg_collection)
