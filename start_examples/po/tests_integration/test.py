import os
import unittest

from openpyxl import load_workbook
from start_examples.po.parsers_and_writers.po_writer import PoWriter

from start_examples.po.msg_model.msg_collection import MsgCollection
from start_examples.po.parsers_and_writers.excel_parser import RuEnExcelParser
from start_examples.po.parsers_and_writers.excel_writer import RuEnExcelWriter
from start_examples.po.parsers_and_writers.po_parser import PoParser

path = os.path.join(
    'start_examples', 'po', 'tests_integration'
)
excel_path = os.path.join(path, 'generated', 'test_excel.xlsx')
ru_input_po_path = os.path.join(path, 'ru_input_po_path.po')
en_input_po_path = os.path.join(path, 'en_input_po_path.po')
ru_output_po_path = os.path.join(path, 'generated', 'ru_output_po_path.po')
en_output_po_path = os.path.join(path, 'generated', 'en_output_po_path.po')


class TestIntegration(unittest.TestCase):

    """
    
    Исходный po файл на русском языке:
    
        #: ../modules/user/x.js:112
        #: ../modules/user/x.js:300
        msgid "Cake"
        msgstr "Кекс"
        
        #: ../modules/user/x.js:112
        #: ../modules/user/x.js:300
        msgid "Cookie"
        msgid_plural "Cookies"
        msgstr[0] "Печенька"
        msgstr[1] "Печеньки"
        
    
    Исходный po файл на английском языке:
    
        #: ../modules/user/x.js:112
        #: ../modules/user/x.js:300
        msgid "Cake"
        msgstr "Cake"
        
        #: ../modules/user/x.js:112
        #: ../modules/user/x.js:300
        msgid "Cookie"
        msgid_plural "Cookies"
        msgstr[0] "Cookie"
        msgstr[1] "Cookies"
        
    """

    def test(self):
        # Парвсим исходные po файлы
        ru_msg_collection = PoParser.new().parse_file(ru_input_po_path)
        en_msg_collection = PoParser.new().parse_file(en_input_po_path)

        # Записываем исходные po файлы в excel
        RuEnExcelWriter().write(excel_path, ru_msg_collection, en_msg_collection)

        # Заполняем новые значения коллекции в excel
        fill_new_msg_values_to_excel()

        # Считываем исходные po файлы из excel
        parse_result = RuEnExcelParser().parse(excel_path)

        # Записываем коллекции из excel в po файлы
        PoWriter.new().write_file(ru_output_po_path, parse_result.new_ru_msg_collection)
        PoWriter.new().write_file(en_output_po_path, parse_result.new_en_msg_collection)

        # Считывеам результаты из файла
        ru_msg_output_collection = PoParser.new().parse_file(ru_output_po_path)
        en_msg_output_collection = PoParser.new().parse_file(en_output_po_path)

        # Сравниваем результаты с ожижаемыми
        self.assertEqual(get_ru_expected_collection(), ru_msg_output_collection)
        self.assertEqual(get_en_expected_collection(), en_msg_output_collection)


def fill_new_msg_values_to_excel():
    wb = load_workbook(excel_path)
    ws = wb.get_active_sheet()
    write_row(ws, 2, ['Cake', 'Кекс', 'Cake'])
    write_row(ws, 3, ['Cookie\nCookies', 'Печенька\nПеченьки', 'Cookie\nCookies'])
    wb.save(excel_path)


def get_ru_expected_collection():
    result = MsgCollection()
    result.add_msg(id='Cake', str='Кекс', paths=[
        '../modules/user/x.js:112', '../modules/user/x.js:300'
    ])
    result.add_msg_plural(id='Cookie', id_plural='Cookies', strs=[
        'Печенька', 'Печеньки'
    ], paths=[
        '../modules/user/x.js:112', '../modules/user/x.js:300'
    ])
    return result


def get_en_expected_collection():
    result = MsgCollection()
    result.add_msg(id='Cake', str='Cake', paths=[
        '../modules/user/x.js:112', '../modules/user/x.js:300'
    ])
    result.add_msg_plural(id='Cookie', id_plural='Cookies', strs=[
        'Cookie', 'Cookies'
    ], paths=[
        '../modules/user/x.js:112', '../modules/user/x.js:300'
    ])
    return result


def write_row(ws, row_index, values):
    column_indexes = [4, 5, 6]
    for column_index, value in zip(column_indexes, values):
        ws.cell(row=row_index, column=column_index).value = value
