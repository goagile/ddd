from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Alignment
from openpyxl.utils import get_column_letter

from start_examples.po.utils import join_by_new_line


EXCEL_READY = '\n\t\tExcel готов.\n'
CLOSE_EXCEL = '\n\t\t!!! Закройте Excel !!!\n'


class ExcelWriter:

    @classmethod
    def write(cls, excel_path, msg_collection):
        wb = Workbook()
        ws = wb.get_active_sheet()
        alignment = Alignment(wrapText=True, vertical='top', shrink_to_fit=True)

        for row_index, msg in enumerate(msg_collection, 2):
            if msg.is_plural:
                cell_id = ws.cell(row=row_index, column=1)
                cell_id.alignment = alignment
                cell_id.value = join_by_new_line([msg.id, msg.id_plural])

                cell_en = ws.cell(row=row_index, column=2)
                cell_en.alignment = alignment
                cell_en.value = join_by_new_line(msg.strs)

                cell_paths = ws.cell(row=row_index, column=3)
                cell_paths.alignment = alignment
                cell_paths.value = join_by_new_line(msg.paths)
            else:
                cell_id = ws.cell(row=row_index, column=1)
                cell_id.alignment = alignment
                cell_id.value = msg.id

                cell_en = ws.cell(row=row_index, column=2)
                cell_en.alignment = alignment
                cell_en.value = msg.str

                cell_paths = ws.cell(row=row_index, column=3)
                cell_paths.alignment = alignment
                cell_paths.value = join_by_new_line(msg.paths)

            # ws.row_dimensions[row_index].height = 15

        widths = (25, 30, 80, 20, 40)
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(excel_path)


class RuEnExcelWriter:

    def __init__(self):
        self.names = ['Ключ', 'RU', 'EN', 'Новый ключ', 'Новый RU', 'Новый EN', 'Путь']
        self.widths = [40, 40, 40, 40, 40, 40, 120]
        self.align = Alignment(wrapText=True, vertical='top', shrink_to_fit=True)
        self.wb = Workbook()
        self.ws = self.wb.get_active_sheet()

    def write(self, excel_path, ru_msg_collection, en_msg_collection):
        try:
            self.__write_header()
            self.__write_collections(zip(ru_msg_collection, en_msg_collection))
            self.__set_dimensions()
            self.wb.save(excel_path)
            print(EXCEL_READY)
        except PermissionError:
            print(CLOSE_EXCEL)

    def __write_collections(self, collections):
        for row, (ru, en) in enumerate(collections, start=2):
            if ru.is_plural:
                self.__write_values_to(row, values=[
                    join_by_new_line([ru.id, ru.id_plural]),
                    join_by_new_line(ru.strs),
                    '',
                    '',
                    '',
                    '',
                    join_by_new_line(ru.paths)
                ])
            else:
                self.__write_values_to(row, values=[
                    ru.id,
                    ru.str,
                    '',
                    '',
                    '',
                    '',
                    join_by_new_line(ru.paths)
                ])

    def __write_header(self):
        for i, name in enumerate(self.names, start=1):
            self.ws.cell(row=1, column=i).value = name

    def __write_values_to(self, row, values):
        for col, value in enumerate(values, start=1):
            self.__write_row(row, col, value)

    def __write_row(self, row, col, value):
        cell_en = self.ws.cell(row=row, column=col)
        cell_en.alignment = self.align
        cell_en.value = value

    def __set_dimensions(self):
        # ws.row_dimensions[row_index].height = 15
        for col, width in enumerate(self.widths, start=1):
            self.ws.column_dimensions[get_column_letter(col)].width = width
