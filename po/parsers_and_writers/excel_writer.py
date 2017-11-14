from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Alignment
from openpyxl.utils import get_column_letter


EXCEL_READY = '\n\t\tExcel готов.\n'
CLOSE_EXCEL = '\n\t\t!!! Закройте Excel !!!\n'
COLLECTIONS_MUST_BE_SAME_LENS = 'Коллекции должны быть одинаковой длины'


class RuEnExcelWriter:

    def __init__(self):
        self.names = ['Ключ', 'RU', 'EN', 'Новый ключ', 'Новый RU', 'Новый EN', 'Путь']
        self.widths = [40, 40, 40, 40, 40, 40, 120]
        self.align = Alignment(wrapText=True, vertical='top', shrink_to_fit=True)
        self.wb = Workbook()
        self.ws = self.wb.get_active_sheet()

    def write(self, excel_path, ru_msg_collection, en_msg_collection):
        if len(ru_msg_collection) != len(en_msg_collection):
            raise ValueError(COLLECTIONS_MUST_BE_SAME_LENS)
        try:
            self.__write_header()
            for row_index, (ru_msg, en_msg) in enumerate(zip(ru_msg_collection, en_msg_collection), start=2):
                self.__write_ru_msg_to(row_index, ru_msg)
                self.__write_en_msg_to(row_index, en_msg)
            self.__set_dimensions()
            self.wb.save(excel_path)
            print(EXCEL_READY)
        except PermissionError:
            print(CLOSE_EXCEL)

    def __write_ru_msg_to(self, row_index, msg):
        if msg.is_plural:
            self.__write_row(row_index, 1, join_by_new_line([msg.id, msg.id_plural]))
            self.__write_row(row_index, 2, join_by_new_line(msg.strs))
            self.__write_row(row_index, 7, join_by_new_line(msg.paths))
        else:
            self.__write_row(row_index, 1, msg.id)
            self.__write_row(row_index, 2, msg.str)
            self.__write_row(row_index, 7, join_by_new_line(msg.paths))

    def __write_en_msg_to(self, row_index, msg):
        if msg.is_plural:
            self.__write_row(row_index, 3, join_by_new_line(msg.strs))
        else:
            self.__write_row(row_index, 3, msg.str)

    def __write_header(self):
        for i, name in enumerate(self.names, start=1):
            self.ws.cell(row=1, column=i).value = name

    def __write_values_to(self, row, values):
        for col, value in enumerate(values, start=1):
            self.__write_row(row, col, value)

    def __write_row(self, row, col, value):
        cell = self.ws.cell(row=row, column=col)
        cell.alignment = self.align
        cell.value = value

    def __set_dimensions(self):
        # ws.row_dimensions[row_index].height = 15
        for col, width in enumerate(self.widths, start=1):
            self.ws.column_dimensions[get_column_letter(col)].width = width


def join_by_new_line(collection):
    return '\n'.join(p for p in collection)
